"""Payment & Invoice Tracker - Report Generator"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
from tracker import load_payments, flag_overdue

def generate_excel_report(df, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Status"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    fills = {
        'OVERDUE': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'DUE SOON': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'PAID': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
        'PENDING': PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    }
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        flag = row_data.Flag
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if flag in fills:
                cell.fill = fills[flag]
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 3
    summary_ws = wb.create_sheet("Summary")
    summary_data = [
        ["Metric", "Value"],
        ["Total Invoices", len(df)],
        ["Paid", len(df[df['Flag'] == 'PAID'])],
        ["Overdue", len(df[df['Flag'] == 'OVERDUE'])],
        ["Due Soon", len(df[df['Flag'] == 'DUE SOON'])],
        ["Pending", len(df[df['Flag'] == 'PENDING'])],
        ["Total Overdue Amount (Rs.)", f"{df[df['Flag'] == 'OVERDUE']['Amount'].sum():,.2f}"],
    ]
    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.fill = header_fill
                cell.font = header_font
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 20
    wb.save(output_path)
    print(f"Report generated: {output_path}")

if __name__ == "__main__":
    df = load_payments(os.path.join('..', 'data', 'payments.csv'))
    df = flag_overdue(df)
    df['Due_Date'] = df['Due_Date'].dt.strftime('%Y-%m-%d')
    output_path = os.path.join('..', 'output', 'status_report.xlsx')
    generate_excel_report(df, output_path)
